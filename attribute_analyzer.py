import pandas as pd
import numpy as np
from thefuzz import fuzz
import re
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.comments import Comment
import os
import difflib
from typing import Dict, List, Tuple, cast

def normalize_text(text):
    """Enhanced text normalization"""
    if not isinstance(text, str):
        return str(text)
    # Convert to lowercase and remove extra whitespace
    text = text.lower().strip()
    # Remove periods that aren't between numbers (preserve decimals)
    text = re.sub(r'(?<![\d])\.(?![\d])', '', text)
    # Standardize spaces around parentheses
    text = re.sub(r'\s*\(\s*', ' (', text)
    text = re.sub(r'\s*\)\s*', ') ', text)
    # Remove extra spaces
    text = re.sub(r'\s+', ' ', text)
    return text.strip()

def are_case_variants(attr1, attr2):
    """Check if two attributes are identical except for case"""
    # Remove all whitespace and punctuation for comparison
    clean1 = re.sub(r'[^\w\s]', '', attr1).lower()
    clean2 = re.sub(r'[^\w\s]', '', attr2).lower()
    return clean1 == clean2 and attr1 != attr2

def find_similar_attributes(file_path: str, similarity_threshold: float = 80) -> Dict[str, List[Tuple[str, float]]]:
    """
    Find similar attributes in an Excel file using fuzzy string matching.
    
    Args:
        file_path (str): Path to the Excel file containing attributes
        similarity_threshold (float): Minimum similarity threshold (0-100)
        
    Returns:
        Dict[str, List[Tuple[str, float]]]: Dictionary mapping base attributes to lists of 
            tuples containing (similar_attribute, similarity_score)
    """
    # Read the Excel file
    df = pd.read_excel(file_path)
    
    # Get the column containing attributes (assuming it's the first column)
    attributes = df.iloc[:, 0].dropna().unique()
    
    # Dictionary to store groups of similar attributes
    similar_groups = defaultdict(list)
    
    # Compare each pair of attributes
    for i, attr1 in enumerate(attributes):
        base_normalized = normalize_text(attr1)
        
        for attr2 in attributes[i+1:]:
            # Skip if they're exactly the same
            if attr1 == attr2:
                continue
                
            # Check for case variations first
            if are_case_variants(attr1, attr2):
                similar_groups[attr1].append((attr2, 100))
                continue
            
            compare_normalized = normalize_text(attr2)
            
            # Calculate similarity ratio
            similarity = fuzz.ratio(base_normalized, compare_normalized)
            
            # If similar enough but not identical
            if similarity >= similarity_threshold:
                similar_groups[attr1].append((attr2, similarity))

    return similar_groups

def print_similar_groups(similar_groups, min_threshold):
    print("\nPotential similar attributes found:")
    print("=" * 80)
    
    # Create buckets for each percentage point
    percentage_buckets = defaultdict(list)
    total_pairs = 0
    
    # Sort matches into percentage buckets
    for base_attr, matches in similar_groups.items():
        if matches:
            print(f"\nBase attribute: {base_attr}")
            print("Similar to:")
            for match, score in sorted(matches, key=lambda x: (-x[1], x[0])):
                print(f"  - {match} (similarity: {score}%)")
                percentage_buckets[int(score)].append((base_attr, match))
            print("-" * 40)
    
    # Print summary by percentage
    print("\nSummary by Similarity Percentage:")
    print("=" * 80)
    
    grand_total = 0
    for percent in range(100, int(min_threshold)-1, -1):
        matches = percentage_buckets[percent]
        if matches:
            print(f"\n{percent}% Similarity ({len(matches)} pairs):")
            for idx, (base, match) in enumerate(matches, 1):
                print(f"  {idx}. '{base}' ↔ '{match}'")
            grand_total += len(matches)
    
    print("\n" + "=" * 80)
    print(f"Total number of similar pairs found: {grand_total}")
    print("=" * 80)

def find_differences(str1, str2):
    """Find the differences between two strings and return the parts that differ"""
    matcher = difflib.SequenceMatcher(None, str1, str2)
    differences = []
    
    # Get matching blocks
    blocks = matcher.get_matching_blocks()
    
    # Find the differences between matching blocks
    last_end1, last_end2 = 0, 0
    for block in blocks:
        i, j, size = block
        if i > last_end1 or j > last_end2:
            differences.append((
                str1[last_end1:i],
                str2[last_end2:j]
            ))
        last_end1 = i + size
        last_end2 = j + size
    
    return differences

def get_unique_filename(file_path):
    """Generate a unique filename by appending _1, _2, etc. if the file already exists"""
    if not os.path.exists(file_path):
        return file_path
    
    # Split the filename into name and extension
    directory = os.path.dirname(file_path)
    filename = os.path.basename(file_path)
    name, ext = os.path.splitext(filename)
    
    counter = 1
    while True:
        new_filename = os.path.join(directory, f"{name}_{counter}{ext}")
        if not os.path.exists(new_filename):
            return new_filename
        counter += 1

def export_to_excel(similar_groups, min_threshold, input_file_path):
    """
    Export similar attributes to an Excel file with formatting.
    
    Args:
        similar_groups (dict): Dictionary of similar attribute groups
        min_threshold (float): Minimum similarity threshold used
        input_file_path (str): Path to the input file
        
    Returns:
        str: Path to the generated Excel file
    """
    # Create output filename
    base_name = os.path.splitext(os.path.basename(input_file_path))[0]
    output_file = f"similarity_{base_name}_{min_threshold}%.xlsx"
    output_file = get_unique_filename(output_file)

    # Create workbook and select active sheet
    wb: Workbook = Workbook()
    ws = cast(Worksheet, wb.active)  # Properly cast the active worksheet
    ws.title = f"Similarity {min_threshold}%+"
    
    # Define styles
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    gray_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red_font = Font(color="FF0000")
    normal_font = Font(color="000000")
    bold_font = Font(bold=True)
    
    # Add headers
    base_headers = ["Pair ID", "Attribute", "Similarity %", "Differences"]
    action_headers = ["Proposed Catsy Key", "Merge or Keep Separate?", "Pair ID to merge with", "NOTES"]
    all_headers = base_headers + action_headers
    
    for col, header in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = bold_font
        
        # Apply yellow highlighting to action columns
        if header in action_headers:
            cell.fill = yellow_fill
            
        # Add notes to specific columns
        if header == "Proposed Catsy Key":
            comment = Comment("Alex Kamysz: (all lowercase, no spaces, only special characters allowed are underscores _ )", "Attribute Analyzer")
            cell.comment = comment
        elif header == "Pair ID to merge with":
            comment = Comment("please specify if there are multiple pair IDs to merge", "Attribute Analyzer")
            cell.comment = comment
    
    # Collect and sort all pairs
    all_pairs = []
    for base_attr, matches in similar_groups.items():
        for match, score in matches:
            all_pairs.append((base_attr, match, score))
    
    # Sort by similarity percentage (descending)
    all_pairs.sort(key=lambda x: (-x[2], x[0]))
    
    # Write data
    current_row = 2
    pair_id = 1
    
    for base_attr, match, score in all_pairs:
        # Find differences between the pair
        diffs = find_differences(base_attr, match)
        diff_text = " vs ".join([f"{d1} → {d2}" for d1, d2 in diffs if d1 or d2])
        
        # Write first attribute of pair
        fill = white_fill if pair_id % 2 == 0 else gray_fill
        
        def write_attribute_with_highlighting(row, text, differences):
            cell = ws.cell(row=row, column=2, value=text)
            cell.fill = fill
            cell.font = normal_font
            
            if differences:
                comment = Comment(f"Different parts:\n{', '.join([d[0] if row == current_row else d[1] for d in differences])}", "Attribute Analyzer")
                cell.comment = comment
        
        # First row of pair
        ws.cell(row=current_row, column=1, value=pair_id).fill = fill
        write_attribute_with_highlighting(current_row, base_attr, diffs)
        similarity_cell = ws.cell(row=current_row, column=3, value=score/100)
        similarity_cell.fill = fill
        similarity_cell.number_format = '0%'
        ws.cell(row=current_row, column=4, value=diff_text).fill = fill
        
        # Add empty cells with fill color for action columns
        for col in range(5, len(all_headers) + 1):
            ws.cell(row=current_row, column=col).fill = fill
        
        # Second row of pair
        ws.cell(row=current_row + 1, column=1, value=pair_id).fill = fill
        write_attribute_with_highlighting(current_row + 1, match, diffs)
        similarity_cell = ws.cell(row=current_row + 1, column=3, value=score/100)
        similarity_cell.fill = fill
        similarity_cell.number_format = '0%'
        ws.cell(row=current_row + 1, column=4, value=diff_text).fill = fill
        
        # Add empty cells with fill color for action columns
        for col in range(5, len(all_headers) + 1):
            ws.cell(row=current_row + 1, column=col).fill = fill
        
        current_row += 2
        pair_id += 1
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # Freeze the top row
    ws.freeze_panes = 'A2'
    
    # Save workbook
    wb.save(output_file)
    print(f"\nResults exported to: {output_file}")
    return output_file

if __name__ == "__main__":
    file_path: str = "daemar-full-attribute-list-for-analysis.xlsx"
    
    while True:
        try:
            threshold: float = float(input("Enter minimum similarity threshold (0-100): "))
            if 0 <= threshold <= 100:
                break
            print("Please enter a number between 0 and 100")
        except ValueError:
            print("Please enter a valid number")
    
    similar_groups: Dict[str, List[Tuple[str, float]]] = find_similar_attributes(file_path, similarity_threshold=threshold)
    print(f"\nShowing matches with {threshold}% or higher similarity:")
    print_similar_groups(similar_groups, threshold)
    
    # Export to Excel
    output_file = export_to_excel(similar_groups, threshold, file_path)
    print(f"\nResults exported to: {output_file}")
